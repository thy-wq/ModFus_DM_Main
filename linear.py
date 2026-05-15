import argparse
from sklearn.metrics import confusion_matrix
import numpy as np
import yaml
import os
import torch as th
from dataset import get_dataset
from diffsuion1d.denoising_diffusion_pytorch_1d import (Unet1D, GaussianDiffusion1D,
                                                        Trainer1D, Dataset1D)
from runner.schedule import Schedule
from tqdm import tqdm
import torch.nn as nn
import torch.distributed as dist
from torch.optim.lr_scheduler import CosineAnnealingLR
import pandas as pd
from openpyxl import load_workbook,Workbook
from utils import *

def args_and_config():
    parser = argparse.ArgumentParser()
    # 数据集
    parser.add_argument("--dataset",type=str,default="REAL",choices=['REAL','RML2016a','RML2016b',"RML2018","RML2022"])

    parser.add_argument("--diffusion_step",type=int,default=100)
    parser.add_argument("--runner", type=str, default='linear',help="Choose the mode of runner")
    # parser.add_argument("--config", type=str, default='ddpm18.yml',help="Choose the config file")
    # parser.add_argument("--config", type=str, default='ddpm2211.yml',help="Choose the config file")
    # parser.add_argument("--config", type=str, default='ddpm10b.yml',help="Choose the config file")
    parser.add_argument("--config", type=str, default='ddpmreal.yml',help="Choose the config file")

    parser.add_argument("--model", type=str, default='DDPM_piped',
                        choices=['EDM','DDPM_piped'])
    parser.add_argument("--disc", type=str, default='MCLDNN',
                        help="Choose the discriminator model's structure (ResNet18, PyramidNet)")
    parser.add_argument("--method", type=str, default='PNDM4',
                        help="Choose the numerical methods (DDIM, PNDM2, PNDM4, NDM1, NDM4)")
    parser.add_argument("--sample_step", type=int, default=50,
                        help="Control the total generation step")
    parser.add_argument("--device", type=str, default='cuda',
                        help="Choose the device to use")
    parser.add_argument("--image_path", type=str, default='temp/sample',
                        help="Choose the path to save images")
    parser.add_argument("--model_path", type=str,
                        # default='temp-RML2022-DDPM_piped-100step-pred_noise/train/base_multi/[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]',
                        # default='temp-RML2018-DDPM_piped-100step-pred_noise/train/base_multi/[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23]',
                        # default='temp-RML2016a-DDPM_piped-100step-pred_noise/train/base_multi/[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]',
                        # default='temp-RML2016b-DDPM_piped-100step-pred_noise/train/base_multi/[0, 1, 2, 3, 4, 5, 6, 7, 8, 9]',
                        default='temp-REAL-DDPM_piped-100step-pred_noise/train/base_multi/[0, 1, 2, 3, 4, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]',
                        help="Choose the path of diffusion model")
    parser.add_argument("--restart", action="store_true",
                        help="Restart a previous training process")
    parser.add_argument("--train_path", type=str, default='temp/train/base_multi',
                        help="Choose the path to save training status")
    parser.add_argument("--repeat_size", type=int, default=4,
                        help="Set the model's name")
    # 当模型不是从头开始训练
    # args.resume == ""时从头开始训练，否则加载这里面的参数从该参数继续训练
    parser.add_argument("--resume",type=str,default="")

    # pred_v
    parser.add_argument("--pred_type", default="pred_noise",type=str,choices=['pred_v',"pred_noise","pred_x0"])

    parser.add_argument("--snr",type=int,default=12,help="snr dB used to train the model")

    # For linear evaluation
    parser.add_argument('--linearepoch',type=int, default=50,help="分类头微调的轮次")
    parser.add_argument('--linear_lrrate',type=float, default=0.01,help="分类头微调的学习率")
    parser.add_argument('--linear_bs',type=int, default=64,help="微调分类头时的batchsize")
    # parser.add_argument('--timestep',type=int, default=11,help="获取特征输入的是第timestep步加噪的结果")
    parser.add_argument('--blockname',type=str, default='out_6',help="获取特征的Unet层")
    parser.add_argument('--use_amp',type=bool, default=True,help="")
    parser.add_argument('--grid',type=bool, default=True,help="")

    parser.add_argument("--augment", type=str, default='', choices=['', 'rotation'])
    parser.add_argument("--numlabeld", type=int, default=10, choices=[2,10])
    parser.add_argument("--batch_size", type=int, default=128)
    parser.add_argument("--Monte", type=int, default=10, help='蒙特卡洛实验次数')
    args = parser.parse_args()

    rank = os.environ.get("LOCAL_RANK")
    rank = 0 if rank is None else int(rank)
    world_size = os.environ.get("WORLD_SIZE")
    world_size = 1 if world_size is None else int(world_size)
    if world_size >= 2:
        parser.set_defaults(device=th.device(args.device, rank))
        args = parser.parse_args()
        th.cuda.set_device(rank)

    # assert args.model.lower() in args.config
    work_dir = os.getcwd()
    with open(f'{work_dir}/{args.config}', 'r') as f:
        config = yaml.safe_load(f)
    config['Schedule']['diffusion_step'] = args.diffusion_step
    return args, config


def check_config(config):
    # assert config['Dataset']['batch_size'] == config['Train']['batch_size'] // 2
    pass

def check_sheet_exists(file_path, sheet_name):
    wb = load_workbook(file_path)
    return sheet_name in wb.sheetnames

class ClassifierDict(nn.Module):
    def __init__(self, feat_func, time_list, name_list, base_lr, epoch, img_shape, num_classes):
        super(ClassifierDict, self).__init__()
        self.feat_func = feat_func
        self.times = time_list
        self.names = name_list
        self.classifiers = nn.ModuleDict()
        self.optims = {}
        self.schedulers = {}
        self.loss_fn = nn.CrossEntropyLoss()

        for time in self.times:
            feats = self.feat_func.get_feature(torch.zeros(1, *img_shape).to(device), time)
            if self.names is None:
                self.names = list(feats.keys())  # all available names

            for name in self.names: # 为每一层的输出特征分别训练分类器
                key = self.make_key(time, name)

                # layers = nn.Sequential(nn.Linear(feats[name].shape[1], 128),
                #          nn.Linear(128, num_classes))
                layers = nn.Linear(feats[name].shape[1], num_classes)
                layers = layers.to(device)  # On single GPU, no need for DDP
                optimizer = torch.optim.Adam(layers.parameters(), lr=base_lr)
                # optimizer = torch.optim.Adam(list(self.feat_func.model.parameters())+list(layers.parameters()), lr=base_lr)
                scheduler = CosineAnnealingLR(optimizer, epoch)
                self.classifiers[key] = layers
                self.optims[key] = optimizer
                self.schedulers[key] = scheduler

    def train(self, x, y):
        self.classifiers.train()
        for time in self.times: # times 获取特征的步长
            feats = self.feat_func.get_feature(x, time) # 获得指定层的输出
            for name in self.names: # 特征的层
                key = self.make_key(time, name)
                representation = feats[name].detach()
                logit = self.classifiers[key](representation)
                loss = self.loss_fn(logit, y.long())

                self.optims[key].zero_grad()
                loss.backward()
                self.optims[key].step()

    def test(self, x):
        outputs = {}
        with torch.no_grad():
            self.classifiers.eval()
            for time in self.times:
                feats = self.feat_func.get_feature(x, time)
                for name in self.names:
                    key = self.make_key(time, name)
                    representation = feats[name].detach()
                    logit = self.classifiers[key](representation)
                    pred = logit.argmax(dim=-1)
                    outputs[key] = pred
        return outputs

    def make_key(self, t, n):
        return str(t) + '/' + n

    def get_lr(self):
        key = self.make_key(self.times[0], self.names[0])
        optim = self.optims[key]
        return optim.param_groups[0]['lr']

    def schedule_step(self):
        for time in self.times:
            for name in self.names:
                key = self.make_key(time, name)
                self.schedulers[key].step()

def train(opt,model):
    def test():
        preds = {k: [] for k in classifiers.optims.keys()}
        accs = {}
        labels = []

        all_conf = {}
        confusion_matrix_test = np.zeros([num_knownclass,num_knownclass],dtype=np.float32)
        y_true, y_pred = [], []
        step = 0
        for image, label in tqdm(valid_loader):
            image = image.squeeze(1)
            outputs = classifiers.test(image.to(device))
            for key in outputs:
                preds[key].append(outputs[key])
            labels.append(label.to(device))

        for key in preds:
            preds[key] = torch.cat(preds[key])
        label = torch.cat(labels)
        # label = gather_tensor(label)
        for key in preds:
            # pred = gather_tensor(preds[key])
            pred = preds[key]
            accs[key] = (pred == label).sum().item() / len(label)
            confusion_matrix_test = confusion_matrix(np.array(label.cpu()), np.array(pred.cpu()))
            all_conf[key] = confusion_matrix_test
        step += 1
        return accs, all_conf


    yaml_path = opt.config
    use_amp = opt.use_amp
    grid_search = opt.grid
    with open(yaml_path, 'r') as f:
        config = yaml.full_load(f)

    epoch = opt.linearepoch
    batch_size = opt.linear_bs
    base_lr = opt.linear_lrrate
    print(f"当前输入为M_{args.cur_monte} 第{args.time}步加噪的样本")
    curtime = args.time # 用来获取特征的步长list(range(8,250,1))list(range(8,250,1))
    name_list = ['out_1', 'out_2', "out_3", 'out_4', 'out_5', 'out_6', "out_7", 'out_8']
    time_list = [curtime]
    # 存储不同次实验中最优识别准确率的list
    accs_blks = [0] * int(8) #

    args.seed = args.cur_monte
    cur_M = args.cur_monte

    # 加载训练集和测试集 重新初始化分类器
    ori_train_loader, valid_loader,_ = get_dataset(args,config['Dataset'])
    dataset, train_loader = sample_n_per_class(args,ori_train_loader.dataset,args.numlabeld)
    classifiers = ClassifierDict(model, time_list, name_list,
                                 base_lr, epoch, opt.image_shape, opt.classes).to(device)
    for e in range(epoch):
        pbar = tqdm(train_loader)
        for i, (image, label) in enumerate(pbar):
            if i% 50 == 0:
                pbar.set_description("[M_%d/time%d/epoch %d / iter %d]: lr: %.1e" % (cur_M,curtime, e, i, classifiers.get_lr()))
            # 输入image [b,3,128,128]  label [128] 原始数据及其标签
            image = image.squeeze(1)
            classifiers.train(image.to(device), label.to(device))
        classifiers.schedule_step()
        if e >48:
            accs, all_conf = test()
            print("\n")
            print0(accs)
            keys = list(accs.keys())
            # 更新最高识别准确率
            accs_blks[0] = max(accs[keys[0]], accs_blks[0])
            accs_blks[1] = max(accs[keys[1]], accs_blks[1])
            accs_blks[2] = max(accs[keys[2]], accs_blks[2])
            accs_blks[3] = max(accs[keys[3]], accs_blks[3])
            accs_blks[4] = max(accs[keys[4]], accs_blks[4])
            accs_blks[5] = max(accs[keys[5]], accs_blks[5])
            accs_blks[6] = max(accs[keys[6]], accs_blks[6])
            accs_blks[7] = max(accs[keys[7]], accs_blks[7])

    if args.cur_monte == 0:
        df = pd.DataFrame({"block":keys,
                           f"M_{args.cur_monte}":accs_blks})
    else:
        df = pd.DataFrame({f"M_{args.cur_monte}": accs_blks})
    # file_path = os.path.join(args.model_path[:-11],f"{args.snr}dB_{args.numlabeld}","Finetune",f"{str(args.snr)}dB_{str(args.numlabeld)}_linear.xlsx")
    file_path = os.path.join(args.model_path[:-11],f"{str(args.snr)}dB_{str(args.numlabeld)}","SingelBlock",f"{str(args.snr)}dB_{str(args.numlabeld)}_linear.xlsx")
    print(f"save in {file_path}")
    os.makedirs(os.path.join(args.model_path[:-11],f"{str(args.snr)}dB_{str(args.numlabeld)}","SingelBlock"), exist_ok=True)
    if not os.path.exists(file_path): # 如果不存在，则创建一个新的表格
        wb = Workbook()
        wb.save(file_path)
        print(f"当前创建的表路径为{file_path}")
    if check_sheet_exists(file_path, f"time{curtime}"):
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            existing_df = pd.read_excel(file_path, sheet_name=f"time{curtime}")
            combined_df = pd.concat([existing_df, df], axis=1, ignore_index=True)
            writer.book.remove(writer.book[f"time{curtime}"])
            combined_df.to_excel(writer, sheet_name=f"time{curtime}", index=False)
    else:
        with pd.ExcelWriter(file_path,engine='openpyxl',mode='a',if_sheet_exists='overlay') as writer:
            df.to_excel(writer,index=False,sheet_name=f"time{curtime}")
            print(f"当前写入的表名称为time{curtime}")

def print0(*args, **kwargs):
    if 'LOCAL_RANK' not in os.environ or int(os.environ['LOCAL_RANK']) == 0:
        print(*args, **kwargs)

if __name__ == "__main__":
    # 在如args和config配置-
    # N_list = [2, 5, 10]
    all_times = list(range(1, 15, 1))
    Monte = 10
    for curtime in all_times:
        for cur_monte in range(Monte):
            # 在如args和config配置
            args, config = args_and_config()
            check_config(config)
            args.time = curtime
            args.cur_monte = cur_monte
            print(f"当前输入为第m_{args.cur_monte} step{args.time}步加噪的样本")

            if args.dataset == "RML2018":
                args.image_path = os.path.join(args.image_path, str(args.snr) + "dB", "sample")
                args.model_path = os.path.join(args.model_path, str(args.snr) + "dB", 'train.ckpt')
            elif args.dataset == "REAL":
                args.image_path = os.path.join(args.image_path, "24dB", "sample")
                args.model_path = os.path.join(args.model_path, "24dB", 'train.ckpt')
                print(f"模型参数加载路径{args.model_path}")
                print(f"模型参数加载路径{args.model_path}")
            else:
                args.image_path = os.path.join(args.image_path, "100dB", "sample")
                args.model_path = os.path.join(args.model_path, "100dB", 'train.ckpt')
            print(f"模型参数加载路径{args.model_path}")

            device = th.device(args.device)
            schedule = Schedule(args, config['Schedule'])
            diffusion_step = config["Schedule"]["diffusion_step"]
            num_knownclass = len(config['Dataset']['known'])

            assert config['Model']['struc'] == args.model

            # Load diffusion model 加载DM模型
            if config['Model']['struc'] == 'DDPM_piped':
                if args.dataset == "RML2016a" or args.dataset == "RML2016b"or args.dataset == "RML2022":
                    model = Unet1D(dim=64, dim_mults=(1, 2, 4, 8), channels=2, num_classes=num_knownclass,use_hooks=True).cuda()
                    diffusion = GaussianDiffusion1D(model, seq_length=128, timesteps=diffusion_step,
                                                    objective='pred_noise').cuda()
                    args.image_shape = [2,1,128]
                    args.classes = num_knownclass
                elif args.dataset == "RML2018" or args.dataset == "REAL":
                    model = Unet1D(dim=64, dim_mults=(1, 2, 4, 8), channels=2, num_classes=num_knownclass,use_hooks=True).cuda()
                    diffusion = GaussianDiffusion1D(model, seq_length=1024, timesteps=diffusion_step,
                                                    objective='pred_noise').cuda()
                    args.image_shape = [2, 1, 1024]
                    args.classes = num_knownclass
            elif config['Model']['struc'] == 'EDM':
                print("Use EDM diffusion model")
                import thop
                from diffsuion1d.EDM import EDM
                if args.dataset == "RML2016a" or args.dataset == "RML2016b" or args.dataset == "RML2022":
                    assert config['Dataset']['name'] == args.dataset
                    args.image_shape = [2, 1, 128]
                    num_knownclass = len(config['Dataset']['known'])
                    args.classes = num_knownclass
                    model = Unet1D(dim=64, dim_mults=(1, 2, 4, 8), channels=2, num_classes=num_knownclass).cuda()
                    diffusion = EDM(model, seq_length=128, timesteps=diffusion_step,
                                                    objective=args.pred_type).cuda()
            else:
                print("输入正确的扩散模型参数：config['Model']['struc']")
            if args.runner == 'training': # 训练模型
                from runner.runner import Runner
                runner = Runner(args, config, schedule, diffusion)
                runner.train_loop()
                print(f'Do not find the runner {args.runner}.')
            elif args.runner == 'fid': # 生成数据
                from runner.runner_classes import Runner
                runner = Runner(args, config, schedule, diffusion)
                runner.sample_fid()
            elif args.runner == 'linear':
                print("线性评估")
                from runner.runner_classes import Runner
                runner = Runner(args, config, schedule, diffusion)
                runner.pre_get_activation()# 加载模型参数
                train(args, runner)

